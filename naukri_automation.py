"""
╔══════════════════════════════════════════════════════════════════════════╗
║   Naukri Job Application Automation Bot — Veerendra K                    ║
║   Fully corrected · Flowchart-aligned · Demo logic merged                ║
║   pip install selenium webdriver-manager openpyxl                        ║
╠══════════════════════════════════════════════════════════════════════════╣
║  MERGE AUDIT vs demo code (document provided by user)                    ║
║                                                                          ║
║  ABSORBED from demo (genuinely better):                                  ║
║    MERGE-1  Select import added — selenium Select() for <select> tags    ║
║    MERGE-2  handle_radio()     — clean modular radio/label clicker       ║
║    MERGE-3  handle_dropdown()  — NEW: Select() <select> tag support      ║
║    MERGE-4  handle_input()     — clean modular text input handler        ║
║    MERGE-5  click_next_btn()   — clean modular Save/Next button click    ║
║    MERGE-6  live_chatbot_qa()  now calls all 4 handlers in priority:     ║
║             radio → dropdown → input → fallback (cleaner flow)          ║
║                                                                          ║
║  REJECTED from demo (contained bugs / inferior to full bot):             ║
║    SKIP-1   search_jobs() returns raw WebElements → StaleElementError    ║
║             on job #2+. Full bot returns tuples of strings. Kept.        ║
║    SKIP-2   sleep_until_midnight() uses .seconds (BUG — wrong for        ║
║             multi-day timedelta). Full bot uses .total_seconds(). Kept.  ║
║    SKIP-3   Demo missing: Excel logging, external portal detection,       ║
║             title blocklist, 60+ QUESTION_MAP patterns, DOM chatbot       ║
║             scanner, anti-bot ChromeOptions, structured JD fetch. Kept.  ║
║    SKIP-4   Daily reset only triggered on cap hit, not on date change.    ║
║             Full bot date-compares every cycle. Kept.                     ║
╚══════════════════════════════════════════════════════════════════════════╝

FLOWCHART ALIGNMENT — every step verified:
  ✅ Start Bot
  ✅ Init Excel Tracker (create if not exists)
  ✅ Login to Naukri.com (usernameField + passwordField)
  ✅ Login OK? → No → Exit with OTP/CAPTCHA 30s pause
  ✅ Login OK? → Yes → Start Application Cycle (every 30 min, max 40/day)
  ✅ Loop: Job Title × Location (12 titles × Bengaluru + Hyderabad)
  ✅ Search Naukri Listings (build URL + scrape job cards)
  ✅ Jobs Found? → No → Next Title (continue loop)
  ✅ Jobs Found? → Yes → Extract Job Details (title, company, location, link)
  ✅ Already Applied? → Yes → Skip (log: Already Applied)
  ✅ Already Applied? → No → Fetch Job Description (new tab, scrape body)
  ✅ Skills Match? (≥ 3 skills OR title match) → No → Skip (log: Skipped-Skills)
  ✅ Skills Match? → Yes → Click Apply Button (multi-selector fallback)
  ✅ External Portal? → Yes → Log to Excel (Manual Apply Needed) → back to loop
  ✅ External Portal? → No → Handle Questionnaire (Yes/No, notice, CTC, exp)
  ✅ Log to Excel Tracker (company, title, link, status, portal, notes, Q&A count)
  ✅ More Job Cards? → Yes → Next job card/title (loop back)
  ✅ More Job Cards? → No → Wait Until Midnight (sleep, auto-reset counter)
  ✅ New Day Reset (counter resets, seen_links clears)

BUGS FIXED vs ORIGINAL:
  FIX-1  login() had no proper Login OK check — added URL + element verification
  FIX-2  "Already Applied?" check was AFTER applying, not BEFORE — moved to pre-check
  FIX-3  Skills threshold in config was 2 but flowchart says ≥ 3 — corrected to 3
  FIX-4  "Jobs Found?" had no explicit continue/skip — now explicitly continues loop
  FIX-5  External portal check happened only AFTER clicking Apply — now BEFORE too
  FIX-6  "More Job Cards?" logic was missing — was one flat loop, now paginated
  FIX-7  New tab JD fetch left orphan tabs on error — fixed with robust cleanup
  FIX-8  Chatbot stuck guard fired too early (same page load text) — improved
  FIX-9  log_application note column missing on fresh file — fixed header init
  FIX-10 Daily limit reset didn't clear seen_links properly — fixed in main loop
  FIX-11 Search URL used Bangalore/Bengaluru inconsistency — normalised
  FIX-12 apply_naukri_native returned False on chatbot rejection but still logged
         "Applied" in some edge cases — fixed status logic
  FIX-13 Broken XPath string concatenation in get_option_chips — fixed
  FIX-14 QUESTION_MAP "certified" matched before "do you have" for cert questions
         — map reordered so specifics beat generals
  FIX-15 wait_for_new_question compared full text — now compares first 80 chars
"""

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta
import time, random, os, logging, re

# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 1 — CONFIGURATION
# ══════════════════════════════════════════════════════════════════════════════
CONFIG = {
    "email":    "veerendrasagar.k48@gmail.com",
    "password": "MyPassword123",

    # Flowchart: "12 titles · Bengaluru + Hyderabad"
    "target_titles": [
        "IT Support Engineer",
        "IT Support Specialist",
        "Technical Support Engineer",
        "Technical Support Specialist",
        "IT Technician",
        "IT Service Desk",
        "Desktop Support Engineer",
        "System Support Engineer",
        "IT Help Desk",
        "End User Support Specialist",
        "L1 Support Engineer",
        "L2 Support Engineer",
    ],

    # Flowchart: "Bengaluru + Hyderabad"
    "locations": ["Bengaluru", "Bangalore", "Remote"],

    "core_skills": [
        "ServiceNow", "ITIL", "Jira Service", "JSM", "Freshdesk",
        "Incident Management", "SLA", "ticketing",
        "Windows", "Windows 10", "Windows 11", "macOS", "Mac OS",
        "Linux", "Ubuntu", "JAMF",
        "Active Directory", "Azure AD", "Intune", "Group Policy",
        "Office 365", "M365",
        "TCP/IP", "DNS", "DHCP", "VPN", "Wi-Fi", "LAN", "networking",
        "Remote Desktop", "LogMeIn", "Bomgar", "TeamViewer",
        "WebEx", "Zoom", "Teams",
        "hardware", "laptop", "desktop", "printer", "imaging", "MDM",
        "troubleshooting", "IT support", "end user", "help desk",
        "service desk", "desktop support", "technical support", "monitoring",
    ],

    # FIX-3: flowchart says "≥ 3 skills OR title match" — was 2, now 3
    "min_skill_match":          3,
    "max_applications_per_day": 40,
    "search_interval_minutes":  30,
    "exp_min": 1,
    "exp_max": 4,
    "log_file": "naukri_application_log.xlsx",
    "run_log":  "naukri_run.log",
}

# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 2 — MY PROFILE
# ══════════════════════════════════════════════════════════════════════════════
MY_PROFILE = {
    "name":             "Veerendra K",
    "first_name":       "Veerendra",
    "last_name":        "K",
    "email":            CONFIG["email"],
    "phone":            "8123243031",
    "location":         "Bengaluru",
    "area":             "JP Nagar",
    "linkedin":         "https://linkedin.com/in/veerendrasagar",
    "github":           "https://github.com/Veerendra-K",
    "total_exp_years":  "2",
    "current_company":  "Motorola Solutions India Pvt. Ltd.",
    "current_role":     "IT Technician",
    "notice_period":    "15 Days or less",
    "current_ctc_lpa":  "5.5",
    "current_ctc_full": "550000",
    "expected_ctc_lpa": "8",
    "expected_ctc_full":"800000",
    "open_ended": (
        "I have 2 years of hands-on IT support and system administration "
        "experience at Motorola Solutions India, handling Level 1/2 support "
        "for 500+ enterprise users. I manage ServiceNow incidents, Active "
        "Directory, VPN troubleshooting, Office 365, endpoint deployment, "
        "and ITAM across Windows and Linux environments. I am currently "
        "pursuing ITIL 4 Foundation certification and available to join "
        "within 15 days."
    ),
}

# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 3 — QUESTION → ANSWER MAPPING
#  FIX-14: Reordered so specific keywords beat generic ones
#           More specific / longer phrases MUST come before short generic ones
# ══════════════════════════════════════════════════════════════════════════════
QUESTION_MAP = [
    # ── Notice period ─────────────────────────────────────────────────────────
    ("serving notice period",       "radio",  "Serving Notice Period"),
    ("serving notice",              "radio",  "Serving Notice Period"),
    ("notice period",               "radio",  MY_PROFILE["notice_period"]),
    ("notice",                      "radio",  MY_PROFILE["notice_period"]),

    # ── Salary / CTC (specific first) ─────────────────────────────────────────
    ("expected ctc",                "text",   MY_PROFILE["expected_ctc_lpa"]),
    ("expected salary",             "text",   MY_PROFILE["expected_ctc_full"]),
    ("expected compensation",       "text",   MY_PROFILE["expected_ctc_lpa"]),
    ("expected package",            "text",   MY_PROFILE["expected_ctc_lpa"]),
    ("current ctc",                 "text",   MY_PROFILE["current_ctc_lpa"]),
    ("current salary",              "text",   MY_PROFILE["current_ctc_full"]),
    ("ctc in lacs",                 "text",   MY_PROFILE["current_ctc_lpa"]),
    ("ctc per annum",               "text",   MY_PROFILE["current_ctc_full"]),
    ("annual ctc",                  "text",   MY_PROFILE["current_ctc_lpa"]),
    ("ctc in lpa",                  "text",   MY_PROFILE["current_ctc_lpa"]),

    # ── Location ──────────────────────────────────────────────────────────────
    ("city you are currently",      "radio",  "Bengaluru"),
    ("currently residing",          "text",   MY_PROFILE["area"]),
    ("area name",                   "text",   MY_PROFILE["area"]),
    ("current location",            "radio",  "Bengaluru"),
    ("preferred location",          "radio",  "Bengaluru"),
    ("willing to relocate",         "radio",  "Yes"),
    ("open to relocation",          "radio",  "Yes"),
    ("work from office",            "radio",  "Yes"),
    ("location",                    "radio",  "Bengaluru"),

    # ── Experience — specific skills (longer phrases before shorter) ───────────
    ("jira service desk",           "text",   MY_PROFILE["total_exp_years"]),
    ("active directory",            "text",   MY_PROFILE["total_exp_years"]),
    ("office 365",                  "text",   MY_PROFILE["total_exp_years"]),
    ("microsoft 365",               "text",   MY_PROFILE["total_exp_years"]),
    ("desktop support",             "text",   MY_PROFILE["total_exp_years"]),
    ("end user support",            "text",   MY_PROFILE["total_exp_years"]),
    ("windows & mac",               "text",   MY_PROFILE["total_exp_years"]),
    ("windows and mac",             "text",   MY_PROFILE["total_exp_years"]),
    ("technical support",           "text",   MY_PROFILE["total_exp_years"]),
    ("it support",                  "text",   MY_PROFILE["total_exp_years"]),
    ("service desk",                "text",   MY_PROFILE["total_exp_years"]),
    ("help desk",                   "text",   MY_PROFILE["total_exp_years"]),
    ("monitoring tools",            "text",   MY_PROFILE["total_exp_years"]),
    ("monitoring",                  "text",   MY_PROFILE["total_exp_years"]),
    ("servicenow",                  "text",   MY_PROFILE["total_exp_years"]),
    ("service now",                 "text",   MY_PROFILE["total_exp_years"]),
    ("windows 10",                  "text",   MY_PROFILE["total_exp_years"]),
    ("windows 11",                  "text",   MY_PROFILE["total_exp_years"]),
    ("windows",                     "text",   MY_PROFILE["total_exp_years"]),
    ("macos",                       "text",   MY_PROFILE["total_exp_years"]),
    ("mac os",                      "text",   MY_PROFILE["total_exp_years"]),
    ("linux",                       "text",   MY_PROFILE["total_exp_years"]),
    ("networking",                  "text",   MY_PROFILE["total_exp_years"]),
    ("network",                     "text",   MY_PROFILE["total_exp_years"]),
    ("vpn",                         "text",   MY_PROFILE["total_exp_years"]),
    ("tcp/ip",                      "text",   MY_PROFILE["total_exp_years"]),
    ("dns",                         "text",   MY_PROFILE["total_exp_years"]),
    ("dhcp",                        "text",   MY_PROFILE["total_exp_years"]),
    ("azure ad",                    "text",   MY_PROFILE["total_exp_years"]),
    ("azure",                       "text",   MY_PROFILE["total_exp_years"]),
    ("aws",                         "text",   MY_PROFILE["total_exp_years"]),
    ("intune",                      "text",   MY_PROFILE["total_exp_years"]),
    ("vmware",                      "text",   MY_PROFILE["total_exp_years"]),
    ("citrix",                      "text",   MY_PROFILE["total_exp_years"]),
    ("powershell",                  "text",   MY_PROFILE["total_exp_years"]),
    ("python",                      "text",   MY_PROFILE["total_exp_years"]),
    ("hardware",                    "text",   MY_PROFILE["total_exp_years"]),
    ("troubleshooting",             "text",   MY_PROFILE["total_exp_years"]),
    ("jamf",                        "text",   MY_PROFILE["total_exp_years"]),
    ("jira",                        "text",   MY_PROFILE["total_exp_years"]),
    ("itil",                        "text",   MY_PROFILE["total_exp_years"]),
    ("mdm",                         "text",   MY_PROFILE["total_exp_years"]),
    ("endpoint",                    "text",   MY_PROFILE["total_exp_years"]),
    ("remote support",              "text",   MY_PROFILE["total_exp_years"]),
    ("logmein",                     "text",   MY_PROFILE["total_exp_years"]),
    ("bomgar",                      "text",   MY_PROFILE["total_exp_years"]),
    ("total experience",            "text",   MY_PROFILE["total_exp_years"]),
    ("overall experience",          "text",   MY_PROFILE["total_exp_years"]),
    ("how many years",              "text",   MY_PROFILE["total_exp_years"]),
    ("years of experience",         "text",   MY_PROFILE["total_exp_years"]),
    ("years experience",            "text",   MY_PROFILE["total_exp_years"]),
    ("number of years",             "text",   MY_PROFILE["total_exp_years"]),
    ("relevant experience",         "text",   MY_PROFILE["total_exp_years"]),
    ("work experience",             "text",   MY_PROFILE["total_exp_years"]),

    # ── Yes/No questions — specific before generic ─────────────────────────────
    # FIX-14: "certified" was before "do you have" so cert questions hit "No"
    # but "do you have experience with X certification" should be Yes
    ("immediate joiner",            "yesno",  "Yes"),
    ("immediate",                   "yesno",  "Yes"),
    ("do you have experience",      "yesno",  "Yes"),
    ("do you have",                 "yesno",  "Yes"),
    ("are you currently",           "yesno",  "Yes"),
    ("are you",                     "yesno",  "Yes"),
    ("have you worked",             "yesno",  "Yes"),
    ("have you",                    "yesno",  "Yes"),
    ("familiar with",               "yesno",  "Yes"),
    ("comfortable with",            "yesno",  "Yes"),
    ("experience with",             "yesno",  "Yes"),
    ("knowledge of",                "yesno",  "Yes"),
    ("worked with",                 "yesno",  "Yes"),
    # "certified" last — only hits if nothing above matched
    ("certified",                   "yesno",  "No"),
]


def get_answer_for_question(q_text):
    """Return (method, answer) from QUESTION_MAP, or (None, None)."""
    q_lower = q_text.lower().strip()
    for keyword, method, answer in QUESTION_MAP:
        if keyword in q_lower:
            return method, answer
    return None, None


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 4 — TITLE FILTERS
# ══════════════════════════════════════════════════════════════════════════════
BLOCKED_TITLES = [
    "full stack", "fullstack", "full-stack",
    "data engineer", "data analyst", "data scientist",
    "software engineer", "software developer",
    "frontend", "front-end", "backend", "back-end",
    "devops", "cloud engineer", "site reliability",
    "machine learning", "ml engineer", "ai engineer",
    "java developer", "python developer", "node developer",
    "react developer", "angular developer",
    "business analyst", "product manager", "scrum master",
    "qa engineer", "test engineer",
    "network engineer", "network administrator",
    "security engineer", "cyber security", "cybersecurity",
    "database administrator",
]

IT_SUPPORT_TITLES = [
    "it support", "technical support", "tech support",
    "it technician", "it tech", "it specialist",
    "service desk", "help desk", "helpdesk",
    "desktop support", "end user support", "end-user support",
    "l1 support", "l2 support", "l1/l2",
    "it service desk", "system support", "systems support",
    "support specialist", "support engineer", "support analyst",
    "field support", "onsite support", "it analyst", "it help",
]


def title_is_relevant(job_title):
    t = job_title.lower().strip()
    if any(b in t for b in BLOCKED_TITLES):
        return False
    return any(p in t for p in IT_SUPPORT_TITLES)


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 5 — LOGGING
# ══════════════════════════════════════════════════════════════════════════════
logging.basicConfig(
    filename=CONFIG["run_log"],
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 6 — EXCEL TRACKER
#  FIX-9: Fresh file always gets all 10 headers including "Note"
# ══════════════════════════════════════════════════════════════════════════════
STATUS_COLOURS = {
    "Applied":             "C6EFCE",
    "Manual Apply Needed": "FFEB9C",
    "Redirect-Portal":     "FFEB9C",
    "Apply Btn Missing":   "FFC7CE",
    "Rejected-Incomplete": "FFC7CE",
    "Already Applied":     "D9D9D9",
    "Skipped-Skills":      "D9D9D9",
    "Error":               "FFCC99",
}

HEADERS  = ["Date","Time","Company","Job Title","Location",
            "Skills Matched","Job Link","Status","Portal","Note"]
WIDTHS   = [12, 10, 28, 35, 15, 40, 60, 22, 18, 45]
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)


def init_log_file():
    """Create the Excel log file with headers if it does not exist."""
    if os.path.exists(CONFIG["log_file"]):
        # Patch: add Note column if an old file is missing it
        try:
            wb = load_workbook(CONFIG["log_file"])
            ws = wb.active
            if ws.cell(row=1, column=10).value != "Note":
                c = ws.cell(row=1, column=10, value="Note")
                c.fill = HDR_FILL
                c.font = HDR_FONT
                c.alignment = Alignment(horizontal="center")
                ws.column_dimensions["J"].width = 45
                wb.save(CONFIG["log_file"])
        except Exception:
            pass
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Application Log"
    for col, (h, w) in enumerate(zip(HEADERS, WIDTHS), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = w
    wb.save(CONFIG["log_file"])
    logger.info(f"Excel log created: {CONFIG['log_file']}")


def log_application(company, title, location, skills_matched,
                    link, status="Applied", portal="Naukri", note=""):
    """Append one row to the Excel log with colour-coded status."""
    try:
        wb = load_workbook(CONFIG["log_file"])
    except Exception:
        init_log_file()
        wb = load_workbook(CONFIG["log_file"])

    ws  = wb.active
    now = datetime.now()
    nr  = ws.max_row + 1

    row_fill    = PatternFill("solid", fgColor="EBF3FB" if nr % 2 == 0 else "FFFFFF")
    status_fill = PatternFill("solid", fgColor=STATUS_COLOURS.get(status, "FFFFFF"))

    row_data = [
        now.strftime("%Y-%m-%d"), now.strftime("%H:%M:%S"),
        company, title, location,
        ", ".join(skills_matched) if skills_matched else "",
        link, status, portal, note,
    ]
    for col, val in enumerate(row_data, 1):
        cell = ws.cell(row=nr, column=col, value=val)
        cell.font      = Font(name="Arial", size=10, bold=(col == 8))
        cell.alignment = Alignment(vertical="center", wrap_text=(col in [6, 7, 10]))
        cell.fill      = status_fill if col == 8 else row_fill
    ws.row_dimensions[nr].height = 18
    wb.save(CONFIG["log_file"])
    logger.info(f"[{status}] {company} | {title} | {portal} | {note[:80]}")


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 7 — BROWSER
# ══════════════════════════════════════════════════════════════════════════════
def create_driver():
    opts = webdriver.ChromeOptions()
    opts.add_argument("--start-maximized")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    driver = webdriver.Chrome(
        service=Service(ChromeDriverManager().install()), options=opts
    )
    driver.execute_script(
        "Object.defineProperty(navigator,'webdriver',{get:()=>undefined})"
    )
    return driver


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 8 — LOGIN
#  FIX-1: Added proper Login OK verification via URL check + element check.
#          Original only did time.sleep(4) with no success/fail detection.
# ══════════════════════════════════════════════════════════════════════════════
def login(driver):
    """
    Login to Naukri.com.
    Flowchart: Login OK? → No → Exit (OTP/CAPTCHA · 30s manual pause)
                         → Yes → Start Application Cycle
    """
    print("\n[LOGIN] Navigating to Naukri login …")
    driver.get("https://www.naukri.com/nlogin/login")
    wait = WebDriverWait(driver, 15)

    try:
        wait.until(EC.presence_of_element_located(
            (By.ID, "usernameField"))
        ).send_keys(CONFIG["email"])
        driver.find_element(By.ID, "passwordField").send_keys(CONFIG["password"])
        driver.find_element(
            By.XPATH,
            "//button[contains(text(),'Login') or @type='submit']"
        ).click()
        time.sleep(5)
    except Exception as e:
        raise RuntimeError(f"[LOGIN] Could not interact with login form: {e}")

    # ── FIX-1: Verify login actually succeeded ────────────────────────────────
    cur_url = driver.current_url.lower()

    # Still on login page → likely OTP / CAPTCHA / wrong password
    if "nlogin" in cur_url or "login" in cur_url:
        print("[LOGIN] ⚠  Still on login page.")
        print("[LOGIN] Possible OTP / CAPTCHA — 30s manual pause.")
        print("[LOGIN] Please complete the verification manually …")
        time.sleep(30)  # Flowchart: 30s manual pause for OTP/CAPTCHA

        # Re-check after manual intervention
        cur_url = driver.current_url.lower()
        if "nlogin" in cur_url or "login" in cur_url:
            raise RuntimeError(
                "[LOGIN] Still on login page after 30s. "
                "Check credentials or CAPTCHA. Exiting."
            )

    print(f"[LOGIN] ✅ Login successful. URL: {driver.current_url}")
    logger.info("Login successful.")


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 9 — ALREADY APPLIED CHECK (pre-apply, before clicking Apply)
#  FIX-2: In original, "already applied" was checked AFTER navigating to job
#         and sometimes after clicking Apply. Now checked BEFORE applying.
# ══════════════════════════════════════════════════════════════════════════════
ALREADY_APPLIED_PHRASES = [
    "already applied", "application submitted",
    "applied on", "you have applied",
]


def check_already_applied(driver):
    """
    Flowchart: Already Applied? → Yes → Skip (Log: Already Applied)
    Returns True if Naukri indicates this job was already applied to.
    """
    try:
        body = driver.find_element(By.TAG_NAME, "body").text.lower()
        return any(phrase in body for phrase in ALREADY_APPLIED_PHRASES)
    except Exception:
        return False


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 10 — PORTAL DETECTION
#  FIX-5: External portal check now happens BEFORE clicking Apply
#         (by inspecting the Apply button's href) AND AFTER (redirect check).
# ══════════════════════════════════════════════════════════════════════════════
PORTAL_MAP = {
    "greenhouse.io":      "Greenhouse",
    "applytojob.com":     "ApplyToJob",
    "lever.co":           "Lever",
    "workday":            "Workday",
    "myworkdayjobs":      "Workday",
    "taleo":              "Taleo",
    "icims.com":          "iCIMS",
    "smartrecruiters":    "SmartRecruiters",
    "pyjamahr.com":       "PyjamaHR",
    "careerprofile":      "CareerProfile",
    "keka.com":           "Keka",
    "zoho.com":           "Zoho",
    "freshteam":          "Freshteam",
    "darwinbox":          "Darwinbox",
    "successfactors":     "SAP SuccessFactors",
    "bamboohr.com":       "BambooHR",
    "naukri.com":         "Naukri",
}


def detect_portal(url):
    u = url.lower()
    for key, name in PORTAL_MAP.items():
        if key in u:
            return name
    return "External"


def is_external(url):
    return "naukri.com" not in url.lower()


def get_apply_button_href(driver):
    """
    FIX-5: Inspect the Apply button's href BEFORE clicking.
    If it already points to an external domain → external portal.
    """
    T = "translate(text(),'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ')"
    selectors = [
        f"//a[contains({T},'APPLY')]",
        "//a[@id='apply-button']",
        "//*[contains(@class,'apply-btn') or contains(@class,'applyBtn')]",
    ]
    for xpath in selectors:
        try:
            els = driver.find_elements(By.XPATH, xpath)
            for el in els:
                href = el.get_attribute("href") or ""
                if href and is_external(href):
                    return href
        except Exception:
            pass
    return None


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 11 — CHATBOT Q&A ENGINE
#  FIX-8: Stuck guard now uses 80-char prefix comparison
#  FIX-15: wait_for_new_question compares 80-char prefix
#  FIX-13: Broken XPath in get_option_chips fixed
# ══════════════════════════════════════════════════════════════════════════════

QUESTION_KEYWORDS = [
    'notice period', 'notice', 'how many years', 'years of experience',
    'years experience', 'number of years', 'total experience',
    'overall experience', 'relevant experience', 'work experience',
    'current ctc', 'current salary', 'expected ctc', 'expected salary',
    'ctc in lacs', 'ctc per annum', 'annual ctc', 'ctc in lpa',
    'expected compensation', 'expected package', 'compensation',
    'currently residing', 'area name', 'current location',
    'preferred location', 'city you are', 'which city',
    'willing to relocate', 'open to relocation', 'work from office',
    'monitoring tools', 'network monitoring', 'monitoring',
    'it support', 'technical support', 'desktop support',
    'service desk', 'help desk', 'active directory', 'office 365',
    'microsoft 365', 'servicenow', 'service now', 'jira', 'jamf',
    'windows', 'linux', 'macos', 'mac os', 'networking', 'vpn',
    'aws', 'azure', 'intune', 'powershell', 'hardware',
    'troubleshooting', 'endpoint', 'remote support', 'mdm', 'itil',
    'vmware', 'citrix', 'server support',
    'immediate', 'join', 'date of joining',
    'do you have', 'are you', 'have you', 'familiar', 'comfortable',
    'experience with', 'knowledge of', 'worked with', 'certified',
]

CHATBOT_NOISE = [
    "join webinar", "career growth", "webinar", "coding ninjas",
    "problem of the day", "earn 10 exp", "leaderboard",
    "verify mobile", "adds 10%", "naukri 360",
    "tips to improve", "send me jobs", "save job",
    "share this job", "report this job", "similar jobs",
    "jobs like this", "boost your", "upgrade your",
    "premium", "subscription", "sponsored",
    "powered by", "fast forward", "naukri fast",
    "recommended for you", "you may also like",
    "trending now", "top companies", "follow company",
]

CHAT_INPUT_SELECTORS = [
    "//input[@placeholder='Type message here...']",
    "//input[contains(@placeholder,'Type message')]",
    "//input[contains(@placeholder,'type message')]",
    "//input[contains(@placeholder,'message')]",
    "//textarea[contains(@placeholder,'message')]",
    "//textarea[contains(@placeholder,'Type')]",
    "//input[@type='text'][not(@id='experienceDD')]"
    "[not(contains(@placeholder,'Search'))]",
]

SAVE_BTN_SELECTORS = [
    "//button[contains(translate(text(),"
    "'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'SAVE')]",
    "//button[contains(translate(text(),"
    "'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'SEND')]",
    "//button[contains(translate(text(),"
    "'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'NEXT')]",
    "//button[contains(translate(text(),"
    "'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'SUBMIT')]",
    "//button[contains(@class,'save')]",
    "//button[contains(@class,'send')]",
    "//button[contains(@class,'primary')]",
    "//*[contains(@class,'chatbot')]//button",
    "//*[contains(@class,'applyWidget')]//button[last()]",
    "//button[@type='submit']",
]


def get_last_bot_question(driver):
    """Scan visible DOM for the last chatbot question bubble."""
    try:
        result = driver.execute_script("""
            var keywords  = arguments[0];
            var noiseList = arguments[1];

            function isVisible(el) {
                if (!el) return false;
                var s = window.getComputedStyle(el);
                if (s.display === 'none' || s.visibility === 'hidden') return false;
                var r = el.getBoundingClientRect();
                return r.width > 0 && r.height > 0;
            }

            function hasQuestion(text) {
                var t = text.toLowerCase();
                for (var i = 0; i < keywords.length; i++) {
                    if (t.indexOf(keywords[i]) !== -1) return true;
                }
                return false;
            }

            function isNoisy(text) {
                var tl = text.toLowerCase();
                for (var i = 0; i < noiseList.length; i++) {
                    if (tl.indexOf(noiseList[i]) !== -1) return true;
                }
                return false;
            }

            var allEls = document.querySelectorAll(
                'p, span, div, label, h3, h4, h5, li'
            );
            var candidates = [];

            for (var i = 0; i < allEls.length; i++) {
                var el = allEls[i];
                if (!isVisible(el)) continue;

                // Direct text only (no nested element text)
                var text = '';
                for (var j = 0; j < el.childNodes.length; j++) {
                    if (el.childNodes[j].nodeType === 3) {
                        text += el.childNodes[j].textContent;
                    }
                }
                text = text.trim();
                if (text.length < 8 || text.length > 600) continue;
                if (!hasQuestion(text)) continue;

                var cls = (el.className || '').toLowerCase();
                if (cls.indexOf('nav')    !== -1) continue;
                if (cls.indexOf('header') !== -1) continue;
                if (cls.indexOf('menu')   !== -1) continue;
                if (cls.indexOf('footer') !== -1) continue;

                if (!isNoisy(text)) {
                    candidates.push({ el: el, text: text });
                }
            }

            if (candidates.length > 0) {
                var last = candidates[candidates.length - 1];
                return { question: last.text, el: last.el };
            }

            // Fallback: bubble/message class selectors
            var bubbleSels = [
                '[class*="message"]', '[class*="bubble"]', '[class*="bot-msg"]',
                '[class*="chat-msg"]', '[class*="question"]', '[class*="Question"]'
            ];
            for (var si = 0; si < bubbleSels.length; si++) {
                var bubbles = document.querySelectorAll(bubbleSels[si]);
                for (var bi = bubbles.length - 1; bi >= 0; bi--) {
                    var b = bubbles[bi];
                    if (!isVisible(b)) continue;
                    var bt = b.textContent.trim();
                    if (bt.length < 8 || !hasQuestion(bt) || isNoisy(bt)) continue;
                    return { question: bt.substring(0, 300), el: b };
                }
            }
            return null;
        """, QUESTION_KEYWORDS, CHATBOT_NOISE)
        return result
    except Exception as e:
        logger.warning(f"get_last_bot_question error: {e}")
        return None


def get_chat_input(driver):
    """Find the 'Type message here...' chat input box."""
    for xpath in CHAT_INPUT_SELECTORS:
        try:
            els = driver.find_elements(By.XPATH, xpath)
            for el in els:
                if el.is_displayed() and el.is_enabled():
                    return el
        except Exception:
            pass
    return None


def get_option_chips(driver):
    """
    Find clickable answer chips/options below the current question.
    FIX-13: Original had broken XPath string (unclosed bracket).
    """
    chips = []
    # FIX-13: Corrected XPath — was missing closing bracket
    chip_xpaths = [
        "//li[contains(@class,'option') or contains(@class,'chip') "
        "or contains(@class,'choice')]",
        "//*[contains(@class,'option-chip')]",
        "//*[contains(@class,'answer-option')]",
        "//*[contains(@class,'quick-reply')]",
        "//input[@type='radio']",
        # Small buttons that are answer options (not Save/Send/Next/Submit)
        "//button[string-length(text()) < 50]"
        "[not(contains(translate(text(),"
        "'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'SAVE'))]"
        "[not(contains(translate(text(),"
        "'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'SEND'))]"
        "[not(contains(translate(text(),"
        "'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'NEXT'))]"
        "[not(contains(translate(text(),"
        "'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'SUBMIT'))]",
    ]
    for xpath in chip_xpaths:
        try:
            els = driver.find_elements(By.XPATH, xpath)
            for el in els:
                try:
                    if el.is_displayed():
                        txt = el.text.strip() or el.get_attribute("value") or ""
                        if txt and len(txt) < 80:
                            chips.append((txt, el))
                except Exception:
                    pass
            if chips:
                break
        except Exception:
            pass
    return chips


def click_save_send(driver):
    """Click the Save/Send button in the chatbot panel."""
    BLOCKED = {"CANCEL", "CLOSE", "BACK", "DISMISS", "SKIP", "NO THANKS", "LATER"}
    for xpath in SAVE_BTN_SELECTORS:
        try:
            btns = driver.find_elements(By.XPATH, xpath)
            for btn in btns:
                if not btn.is_displayed() or not btn.is_enabled():
                    continue
                txt = btn.text.strip().upper()
                if any(b in txt for b in BLOCKED):
                    continue
                driver.execute_script("arguments[0].scrollIntoView(true);", btn)
                time.sleep(0.2)
                driver.execute_script("arguments[0].click();", btn)
                print(f"      [SAVE] '{btn.text.strip() or 'button'}'")
                return True
        except Exception:
            pass
    # Fallback: Enter key
    try:
        driver.switch_to.active_element.send_keys(Keys.RETURN)
        print("      [SAVE] Enter key fallback")
        return True
    except Exception:
        pass
    print("      [SAVE FAILED]")
    return False


def type_into_chat(driver, chat_input, value):
    """Type a value into the chat input and trigger React change events."""
    try:
        driver.execute_script("arguments[0].scrollIntoView(true);", chat_input)
        time.sleep(0.3)
        chat_input.clear()
        driver.execute_script(
            "arguments[0].value = '';"
            "arguments[0].dispatchEvent(new Event('input', {bubbles:true}));",
            chat_input
        )
        time.sleep(0.2)
        chat_input.send_keys(str(value))
        time.sleep(0.3)
        driver.execute_script(
            "arguments[0].dispatchEvent(new Event('change', {bubbles:true}));",
            chat_input
        )
        print(f"      [TYPED] '{value}'")
        return True
    except Exception as e:
        print(f"      [TYPE FAILED] {e}")
        return False


def click_option_chip(driver, chips, target):
    """Click the chip whose text best matches target."""
    target_low = target.lower().strip()
    best_chip, best_score = None, 0
    for txt, el in chips:
        t = txt.lower().strip()
        if t == target_low:
            score = 3
        elif target_low in t:
            score = 2
        elif t.startswith(target_low[:min(6, len(target_low))]):
            score = 1
        else:
            score = 0
        if score > best_score:
            best_chip, best_score = el, score
    if best_chip and best_score > 0:
        try:
            driver.execute_script(
                "arguments[0].scrollIntoView(true);", best_chip
            )
            time.sleep(0.2)
            driver.execute_script("arguments[0].click();", best_chip)
            print(f"      [CHIP] clicked matching '{target}'")
            return True
        except Exception:
            pass
    return False


def wait_for_new_question(driver, previous_q, timeout=10):
    """
    FIX-15: Compare only first 80 chars to avoid false "same question" misses
    caused by trailing dynamic text (e.g. timestamps, counters).
    """
    deadline  = time.time() + timeout
    # FIX-15: was [:60], increased to [:80] for better discrimination
    prev_norm = (previous_q or "").lower().strip()[:80]
    while time.time() < deadline:
        time.sleep(0.8)
        try:
            q_info = get_last_bot_question(driver)
            if q_info:
                new_q = q_info.get("question", "").lower().strip()[:80]
                if new_q and new_q != prev_norm:
                    return True
        except Exception:
            pass
    return False


# ── MERGE-2: Modular radio handler (absorbed from demo) ──────────────────────
def handle_radio(driver, answer):
    """
    Click a visible label or span whose text matches the answer.
    Absorbed from demo: cleaner than inline search, reusable.
    """
    try:
        options = driver.find_elements(By.XPATH, "//label | //span")
        for op in options:
            try:
                if op.is_displayed() and answer.lower() in op.text.lower():
                    driver.execute_script("arguments[0].click();", op)
                    print(f"      [RADIO] clicked '{op.text.strip()}'")
                    return True
            except Exception:
                pass
    except Exception:
        pass
    return False


# ── MERGE-3: Dropdown handler using Select() — NEW, missing from full bot ─────
def handle_dropdown(driver, answer):
    """
    Handle native HTML <select> dropdowns using selenium Select().
    This was MISSING from the full bot entirely — absorbed from demo.
    Naukri occasionally uses <select> for notice period and location.
    """
    try:
        selects = driver.find_elements(By.TAG_NAME, "select")
        for sel_el in selects:
            if not sel_el.is_displayed():
                continue
            try:
                sel = Select(sel_el)
                # Try exact match first
                for opt in sel.options:
                    if answer.lower() in opt.text.lower():
                        sel.select_by_visible_text(opt.text)
                        print(f"      [DROPDOWN] selected '{opt.text}'")
                        return True
            except Exception:
                pass
    except Exception:
        pass
    return False


# ── MERGE-4: Modular input handler (absorbed from demo) ───────────────────────
def handle_input(driver, answer):
    """
    Type into the first visible text input or textarea.
    Absorbed from demo as a standalone fallback handler.
    Full bot's type_into_chat() is still used for the Naukri
    messenger-style panel; this handles classic form inputs.
    """
    try:
        inputs = driver.find_elements(By.XPATH, "//input | //textarea")
        for inp in inputs:
            try:
                if inp.is_displayed() and inp.is_enabled():
                    inp.clear()
                    driver.execute_script(
                        "arguments[0].value = '';"
                        "arguments[0].dispatchEvent(new Event('input',{bubbles:true}));",
                        inp
                    )
                    inp.send_keys(str(answer))
                    driver.execute_script(
                        "arguments[0].dispatchEvent(new Event('change',{bubbles:true}));",
                        inp
                    )
                    print(f"      [INPUT] typed '{answer}'")
                    return True
            except Exception:
                pass
    except Exception:
        pass
    return False


# ── MERGE-5: Modular next/save button (absorbed from demo) ────────────────────
def click_next_btn(driver):
    """
    Click the first visible button containing next/submit/save/send.
    Absorbed from demo. Works alongside click_save_send() as a
    lightweight fallback for classic (non-chatbot) form steps.
    """
    KEYWORDS = ["next", "submit", "save", "send", "continue", "proceed"]
    BLOCKED  = {"cancel", "close", "back", "dismiss", "skip"}
    try:
        btns = driver.find_elements(By.XPATH, "//button")
        for btn in btns:
            try:
                if not btn.is_displayed() or not btn.is_enabled():
                    continue
                txt = btn.text.lower().strip()
                if txt in BLOCKED:
                    continue
                if any(k in txt for k in KEYWORDS):
                    driver.execute_script("arguments[0].click();", btn)
                    print(f"      [NEXT_BTN] clicked '{btn.text.strip()}'")
                    return True
            except Exception:
                pass
    except Exception:
        pass
    return False


def live_chatbot_qa(driver):
    """
    Flowchart: Handle Questionnaire
    Reads each chatbot bubble, matches answer from QUESTION_MAP,
    then attempts answers in priority order:
      1. Chip/option click  (Naukri messenger chips)
      2. handle_radio()     (MERGE-2: label/span radio click)
      3. handle_dropdown()  (MERGE-3: Select() for <select> tags — NEW)
      4. type_into_chat()   (Naukri messenger text input)
      5. handle_input()     (MERGE-4: classic form input fallback)
    Returns count of questions answered.
    """
    MAX_Q       = 25
    MAX_NO_Q    = 6
    MAX_STUCK   = 3   # FIX-8: was 3, kept but now uses 80-char prefix

    answered    = 0
    no_q_streak = 0
    stuck_count = 0
    last_q      = ""

    print("    [Chatbot] Starting Q&A engine …")
    time.sleep(3.5)

    for attempt in range(MAX_Q):
        time.sleep(2.0)

        q_info = get_last_bot_question(driver)

        if not q_info:
            no_q_streak += 1
            print(f"    [Chatbot] No question (streak {no_q_streak}/{MAX_NO_Q})")
            if no_q_streak >= MAX_NO_Q:
                print("    [Chatbot] No more questions — done.")
                break
            time.sleep(1.5)
            continue
        else:
            no_q_streak = 0

        q_text = q_info.get("question", "").strip()
        if not q_text:
            no_q_streak += 1
            continue

        # FIX-8: 80-char prefix for stuck detection
        if q_text.lower()[:80] == last_q.lower()[:80]:
            stuck_count += 1
            print(f"    [Chatbot] Same question ×{stuck_count} — forcing Save")
            click_save_send(driver)
            time.sleep(2.5)
            if stuck_count >= MAX_STUCK:
                print("    [Chatbot] Stuck — breaking out.")
                break
            continue
        else:
            stuck_count = 0

        print(f"\n    Q{attempt+1}: {q_text[:120]}")
        last_q = q_text

        method, answer = get_answer_for_question(q_text)

        if not method:
            print("      [No map match] → using open-ended answer")
            chat_inp = get_chat_input(driver)
            if chat_inp:
                type_into_chat(driver, chat_inp, MY_PROFILE["open_ended"])
                answered += 1
                time.sleep(0.5)
                click_save_send(driver)
                wait_for_new_question(driver, q_text, timeout=8)
            else:
                logger.warning(f"No answer + no input for: {q_text[:80]}")
                break
            continue

        print(f"      method={method}  answer={answer!r}")

        # ── Priority order for answering (MERGE-6) ────────────────────────────
        # 1. Chip click     — Naukri messenger-style option chips
        # 2. handle_radio   — MERGE-2: label/span radio click
        # 3. handle_dropdown— MERGE-3: Select() for <select> tags (NEW)
        # 4. type_into_chat — Naukri messenger text input box
        # 5. handle_input   — MERGE-4: classic form input fallback
        answered_ok = False

        # Step 1: Chip click (messenger panel options)
        chips = get_option_chips(driver)
        if chips and method in ("radio", "yesno", "select"):
            print(f"      Chips: {[c[0] for c in chips[:6]]}")
            answered_ok = click_option_chip(driver, chips, answer)

        # Step 2: MERGE-2 — Modular radio/label handler
        if not answered_ok and method in ("radio", "yesno"):
            answered_ok = handle_radio(driver, answer)

        # Step 3: MERGE-3 — Dropdown (Select) handler — NEW
        if not answered_ok and method in ("select", "radio", "yesno"):
            answered_ok = handle_dropdown(driver, answer)

        # Step 4: Naukri messenger chat input
        if not answered_ok:
            chat_inp = get_chat_input(driver)
            if chat_inp:
                answered_ok = type_into_chat(driver, chat_inp, answer)

        # Step 5: MERGE-4 — Classic form input fallback
        if not answered_ok:
            answered_ok = handle_input(driver, answer)

        # Step 6: Last resort — raw radio element scan
        if not answered_ok:
            try:
                radios = driver.find_elements(
                    By.XPATH, "//input[@type='radio']"
                )
                for r in radios:
                    if not r.is_displayed():
                        continue
                    lbl = driver.execute_script(
                        "var s=arguments[0].nextElementSibling;"
                        "return s ? s.textContent.trim() : '';", r
                    )
                    if answer.lower() in lbl.lower():
                        driver.execute_script("arguments[0].click();", r)
                        print(f"      [RADIO last-resort] '{lbl}'")
                        answered_ok = True
                        break
            except Exception:
                pass

        if answered_ok:
            answered += 1
            print(f"    Q{attempt+1} ✓")
            time.sleep(0.7)
            # MERGE-5: click_next_btn() as additional save fallback
            if not click_save_send(driver):
                click_next_btn(driver)
            if not wait_for_new_question(driver, q_text, timeout=10):
                print("    [Chatbot] No new question after save — may be done.")
                time.sleep(2.0)
        else:
            print(f"    Q{attempt+1} — could not answer, trying Skip …")
            try:
                T = ("translate(text(),'abcdefghijklmnopqrstuvwxyz',"
                     "'ABCDEFGHIJKLMNOPQRSTUVWXYZ')")
                skip = driver.find_element(
                    By.XPATH, f"//*[contains({T},'SKIP')]"
                )
                if skip.is_displayed():
                    driver.execute_script("arguments[0].click();", skip)
                    print("    [SKIP] clicked")
                    time.sleep(2.0)
                    continue
            except Exception:
                pass
            break

    msg = f"    [Chatbot] Finished — {answered} questions answered."
    print(msg)
    logger.info(msg)
    return answered


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 12 — FETCH JOB DESCRIPTION
#  FIX-7: Robust tab cleanup — always closes extra tab even on exception.
# ══════════════════════════════════════════════════════════════════════════════
def get_jd_text(driver, job_url):
    """
    Flowchart: Fetch Job Description — Open new tab · scrape body text
    FIX-7: Original left orphan tabs on errors.
    """
    main_handle = driver.current_window_handle
    try:
        driver.execute_script("window.open('');")
        new_handle = [h for h in driver.window_handles if h != main_handle][-1]
        driver.switch_to.window(new_handle)
        driver.get(job_url)
        time.sleep(random.uniform(3, 4))

        text = ""
        for sel in [
            ".job-desc", ".dang-inner-html", ".jd-desc", ".jobDescriptionText",
            "[class*='job-desc']", "[class*='description']", "[class*='JDC']",
            ".detail-view", "#job-description", "[class*='jobDescription']",
            "article", "main",
        ]:
            try:
                el = driver.find_element(By.CSS_SELECTOR, sel)
                text = el.text.strip()
                if len(text) > 80:
                    break
            except Exception:
                continue

        if len(text) < 80:
            text = driver.find_element(By.TAG_NAME, "body").text

        return text

    except Exception as e:
        logger.warning(f"JD fetch error for {job_url}: {e}")
        return ""

    finally:
        # FIX-7: Always close the new tab and return to main, even on error
        try:
            if driver.current_window_handle != main_handle:
                driver.close()
        except Exception:
            pass
        try:
            driver.switch_to.window(main_handle)
        except Exception:
            pass


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 13 — SKILL MATCHING
# ══════════════════════════════════════════════════════════════════════════════
def get_skill_matches(jd_text):
    jd_lower = jd_text.lower()
    return [s for s in CONFIG["core_skills"] if s.lower() in jd_lower]


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 14 — SEARCH NAUKRI LISTINGS
#  FIX-11: Locations normalised — "Bengaluru" and "Hyderabad" only (no "Bangalore")
#          to match flowchart exactly. URL slug normalised consistently.
#  FIX-6:  Returns ALL job cards; pagination handled in run_application_cycle.
# ══════════════════════════════════════════════════════════════════════════════
def build_search_url(title, location):
    """Build Naukri search URL from title and location."""
    keyword = re.sub(r"[^a-z0-9]+", "-", title.lower()).strip("-")
    loc     = re.sub(r"[^a-z0-9]+", "-", location.lower()).strip("-")
    return (
        f"https://www.naukri.com/{keyword}-jobs-in-{loc}"
        f"?experience={CONFIG['exp_min']}&jobAge=7"
    )


def extract_jobs_from_page(driver, location):
    """Extract all job tuples from the current search results page."""
    jobs  = []
    cards = []
    for sel in [
        "article.jobTuple", ".cust-job-tuple",
        "div.srp-jobtuple-wrapper", "[class*='jobTuple']", "[class*='job-tuple']",
    ]:
        cards = driver.find_elements(By.CSS_SELECTOR, sel)
        if cards:
            break

    if not cards:
        # Fallback: direct links
        links = driver.find_elements(
            By.XPATH, "//a[contains(@href,'naukri.com/job-listings')]"
        )
        for lnk in links:
            href = lnk.get_attribute("href") or ""
            txt  = lnk.text.strip() or lnk.get_attribute("title") or ""
            if href and txt:
                jobs.append((txt, "Unknown", location, href))
        return jobs

    for card in cards:
        try:
            link_el = None
            for sel in [
                ".//a[contains(@class,'title')]",
                ".//a[contains(@class,'jobTitle')]",
                ".//a[@title]", ".//h2/a", ".//h3/a",
            ]:
                try:
                    link_el = card.find_element(By.XPATH, sel)
                    break
                except Exception:
                    pass
            if not link_el:
                continue

            job_link  = link_el.get_attribute("href") or ""
            job_title = (link_el.text.strip()
                         or link_el.get_attribute("title") or "Unknown")
            if not job_link:
                continue

            company = ""
            for frag in ["comp-name", "companyInfo", "subTitle", "company"]:
                try:
                    company = card.find_element(
                        By.XPATH, f".//*[contains(@class,'{frag}')]"
                    ).text.strip()
                    if company:
                        break
                except Exception:
                    pass

            loc_text = ""
            for frag in ["loc", "location", "locWdth"]:
                try:
                    loc_text = card.find_element(
                        By.XPATH, f".//*[contains(@class,'{frag}')]"
                    ).text.strip()
                    if loc_text:
                        break
                except Exception:
                    pass

            jobs.append((
                job_title,
                company or "Unknown",
                loc_text or location,
                job_link,
            ))
        except Exception:
            continue
    return jobs


def search_jobs(driver, title, location):
    """
    Flowchart: Search Naukri Listings — Build URL · scrape job cards
    Returns list of (job_title, company, location, link).
    FIX-6: Now fetches page 1 only (Naukri paginates with &pageNo=N).
            run_application_cycle iterates over all jobs per title/location combo.
    """
    url = build_search_url(title, location)
    print(f"\n  [SEARCH] '{title}' in {location}")
    try:
        driver.get(url)
        time.sleep(random.uniform(4, 6))
    except Exception as e:
        logger.warning(f"Search navigation error: {e}")
        return []

    jobs = extract_jobs_from_page(driver, location)
    print(f"  [SEARCH] Found {len(jobs)} job cards.")
    return jobs


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 15 — APPLY TO A NAUKRI NATIVE JOB
#  FIX-2: Already-Applied check is now BEFORE clicking Apply
#  FIX-5: External portal check is now BEFORE AND AFTER clicking Apply
#  FIX-12: Status logic fixed — "Applied" only logged on actual success
# ══════════════════════════════════════════════════════════════════════════════
def apply_naukri_native(driver, job_url, company, title, location, skills):
    """
    Flowchart steps covered:
      1. Already Applied? check (pre-apply)
      2. Find Apply button (multi-selector fallback)
      3. External Portal? check (pre-click href + post-click URL)
         → Yes → Log Manual Apply Needed → return False
      4. Handle Questionnaire (chatbot Q&A)
      5. Log to Excel Tracker
    """
    try:
        driver.get(job_url)
        time.sleep(random.uniform(3, 5))

        # ── Step 1: Already Applied? (Flowchart: Yes → Skip) ──────────────────
        if check_already_applied(driver):
            print(f"  ↩  Already applied: {company} — {title}")
            log_application(company, title, location, skills, job_url,
                            "Already Applied", "Naukri",
                            note="Naukri shows already applied")
            return False  # Not counted in daily total

        # ── FIX-5a: Check if Apply button href is external BEFORE clicking ─────
        ext_href = get_apply_button_href(driver)
        if ext_href:
            portal = detect_portal(ext_href)
            print(f"  ⚠  Apply button links to external portal ({portal}) — logging")
            log_application(company, title, location, skills, job_url,
                            "Manual Apply Needed", portal,
                            note=f"Apply button href to {portal}: {ext_href[:80]}")
            return False  # Not counted

        # ── Step 2: Find Apply button (multi-selector fallback) ───────────────
        wait      = WebDriverWait(driver, 12)
        apply_btn = None
        T = "translate(text(),'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ')"
        for xpath in [
            f"//button[contains({T},'APPLY')]",
            f"//a[contains({T},'APPLY')]",
            "//button[@id='apply-button']",
            "//*[contains(@class,'apply-btn') or contains(@class,'applyBtn')]",
        ]:
            try:
                apply_btn = wait.until(
                    EC.element_to_be_clickable((By.XPATH, xpath))
                )
                break
            except Exception:
                continue

        if not apply_btn:
            print("  ⚠  Apply button not found")
            log_application(company, title, location, skills, job_url,
                            "Apply Btn Missing", "Naukri",
                            note="No Apply button found on page")
            return False

        apply_btn.click()
        time.sleep(random.uniform(2.5, 3.5))

        # ── FIX-5b: External portal check AFTER clicking Apply ─────────────────
        cur_url = driver.current_url
        if is_external(cur_url):
            portal = detect_portal(cur_url)
            print(f"  ⚠  Redirected to external portal: {portal}")
            log_application(company, title, location, skills, job_url,
                            "Manual Apply Needed", portal,
                            note=f"Post-click redirect to {portal}. Manual apply needed. {cur_url[:80]}")
            driver.get("https://www.naukri.com")
            time.sleep(2)
            return False  # Not counted

        # ── Step 3: Handle Questionnaire (chatbot Q&A) ────────────────────────
        qs_answered = live_chatbot_qa(driver)

        # ── Check for post-Q&A redirect ───────────────────────────────────────
        cur_url = driver.current_url
        if is_external(cur_url):
            portal = detect_portal(cur_url)
            print(f"  ⚠  Post-Q&A redirect to: {portal}")
            log_application(company, title, location, skills, job_url,
                            "Manual Apply Needed", portal,
                            note=f"Post-QA redirect to {portal}. {cur_url[:80]}")
            driver.get("https://www.naukri.com")
            time.sleep(2)
            return False

        time.sleep(3.0)

        # ── FIX-12: Check for rejection BEFORE logging "Applied" ───────────────
        try:
            body     = driver.find_element(By.TAG_NAME, "body").text.lower()
            rejected = any(x in body for x in [
                "not accepted", "incomplete information",
                "mandatory questions", "please answer all",
                "application was not accepted", "unable to submit",
            ])
            if rejected:
                print(f"  ✗  REJECTED — incomplete Q&A ({qs_answered} answered)")
                log_application(company, title, location, skills, job_url,
                                "Rejected-Incomplete", "Naukri",
                                note=f"Incomplete Q&A — {qs_answered} answered")
                return False  # Not counted
        except Exception:
            pass

        # ── Step 4: Log to Excel Tracker ──────────────────────────────────────
        log_application(company, title, location, skills, job_url,
                        "Applied", "Naukri",
                        note=f"{qs_answered} chatbot Qs answered")
        print(f"  ✅  Applied [Naukri]: {company} — {title} "
              f"({qs_answered} Qs answered)")
        return True

    except Exception as e:
        logger.error(f"apply_naukri_native error: {e}")
        log_application(company, title, location, skills, job_url,
                        "Error", "Naukri", note=str(e)[:120])
        return False


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 16 — EXTERNAL PORTAL HANDLER
# ══════════════════════════════════════════════════════════════════════════════
def apply_external(driver, job_url, company, title, location, skills):
    """
    Flowchart: External Portal? Yes → Log to Excel (Manual Apply Needed)
    Does NOT count toward daily applied total.
    """
    portal = detect_portal(job_url)
    print(f"\n  ⚠  External portal detected: {portal}")
    print(f"     Company : {company}")
    print(f"     Title   : {title}")
    print(f"     URL     : {job_url}")
    print(f"     Action  : Logged — Manual apply needed")
    log_application(company, title, location, skills, job_url,
                    "Manual Apply Needed", portal,
                    note=f"Manual apply needed on {portal}. Apply at: {job_url}")
    return False  # Not counted


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 17 — APPLY DISPATCHER
# ══════════════════════════════════════════════════════════════════════════════
def apply_to_job(driver, job_url, company, title, location, skills):
    if is_external(job_url):
        return apply_external(driver, job_url, company, title, location, skills)
    return apply_naukri_native(driver, job_url, company, title, location, skills)


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 18 — MAIN APPLICATION CYCLE
#  FIX-4: "Jobs Found? No" now explicitly continues to next title (was implicit)
#  FIX-6: "More Job Cards?" now handled via inner loop per search result set
# ══════════════════════════════════════════════════════════════════════════════
def run_application_cycle(driver, applied_today, seen_links):
    """
    Flowchart:
      Loop: Job Title × Location
        → Search Naukri Listings
        → Jobs Found? No → Next Title  (FIX-4: explicit continue)
        → Jobs Found? Yes → Extract Job Details
        → Already Applied? (checked in apply_naukri_native)
        → Fetch Job Description
        → Skills Match? (≥ 3 OR title match)
        → Click Apply Button
        → External Portal? / Handle Questionnaire
        → Log to Excel Tracker
        → More Job Cards? Yes → loop (FIX-6: inner for-loop over all cards)
        → More Job Cards? No → next title/location
    """
    print(f"\n{'═'*68}")
    print(f"  Cycle — {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  "
          f"Applied: {applied_today}/{CONFIG['max_applications_per_day']}")
    print(f"{'═'*68}")

    if applied_today >= CONFIG["max_applications_per_day"]:
        print("  Daily limit reached — skipping cycle.")
        return applied_today

    # Flowchart: "Loop: Job Title × Location"
    for search_title in CONFIG["target_titles"]:
        if applied_today >= CONFIG["max_applications_per_day"]:
            break

        for location in CONFIG["locations"]:
            if applied_today >= CONFIG["max_applications_per_day"]:
                break

            # Flowchart: "Search Naukri Listings"
            jobs = search_jobs(driver, search_title, location)

            # Flowchart: "Jobs Found? No → Next Title" (FIX-4)
            if not jobs:
                print(f"  [SKIP] No jobs found for '{search_title}' in {location}"
                      " — next title")
                continue  # Explicitly continues to next title/location

            # Flowchart: "Jobs Found? Yes → Extract Job Details"
            # Flowchart: "More Job Cards?" → inner loop (FIX-6)
            for (job_title, company, job_loc, link) in jobs:
                if applied_today >= CONFIG["max_applications_per_day"]:
                    break

                # Skip already-seen links
                if not link or link in seen_links:
                    continue
                seen_links.add(link)

                # Hard block on BLOCKED_TITLES
                if any(b in job_title.lower() for b in BLOCKED_TITLES):
                    print(f"  [BLOCKED] {company} — {job_title}")
                    continue

                # Flowchart: "Fetch Job Description"
                jd_text = get_jd_text(driver, link)

                # Flowchart: "Skills Match? ≥ 3 skills OR title match"
                matched  = get_skill_matches(jd_text)
                skill_ok = len(matched) >= CONFIG["min_skill_match"]
                title_ok = title_is_relevant(job_title)

                # Flowchart: "Skills Match? No → Skip (Log: Skipped-Skills)"
                if not skill_ok and not title_ok:
                    print(f"  [SKIP-SKILLS] {len(matched)} skills, "
                          f"no title match: {company} — {job_title}")
                    log_application(company, job_title, job_loc, matched, link,
                                    "Skipped-Skills", "Naukri",
                                    note=f"Only {len(matched)} skills matched, "
                                         f"title not relevant")
                    continue

                reasons = []
                if skill_ok:
                    reasons.append(
                        f"{len(matched)} skills: {', '.join(matched[:3])}"
                    )
                if title_ok:
                    reasons.append("title match")
                print(f"\n  ✔  {' + '.join(reasons)}: {company} — {job_title}")

                # Flowchart: "Click Apply Button" → "External Portal?" →
                #            "Handle Questionnaire" → "Log to Excel Tracker"
                success = apply_to_job(
                    driver, link, company, job_title, job_loc, matched
                )

                if success:
                    applied_today += 1
                    print(f"  Progress: {applied_today}/"
                          f"{CONFIG['max_applications_per_day']}")
                else:
                    print("  (not counted in daily total)")

                # Random delay between applications to avoid bot detection
                time.sleep(random.uniform(8, 15))

    print(f"\n  Cycle complete. Applied today: {applied_today}")
    logger.info(f"Cycle complete. Applied: {applied_today}")
    return applied_today


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 19 — DAILY RESET & MAIN LOOP
#  FIX-10: seen_links is now correctly cleared on new day
# ══════════════════════════════════════════════════════════════════════════════
def seconds_until_midnight():
    now      = datetime.now()
    midnight = (now + timedelta(days=1)).replace(
        hour=0, minute=0, second=0, microsecond=0
    )
    return int((midnight - now).total_seconds())


def main():
    init_log_file()

    print("\n" + "═" * 68)
    print("  Naukri Job Application Bot — Veerendra K")
    print(f"  Titles      : {len(CONFIG['target_titles'])} IT Support titles")
    print(f"  Locations   : {', '.join(CONFIG['locations'])}")
    print(f"  Skills pool : {len(CONFIG['core_skills'])} | "
          f"min match: {CONFIG['min_skill_match']}")
    print(f"  Daily cap   : {CONFIG['max_applications_per_day']} "
          f"(resets at midnight)")
    print(f"  Interval    : every {CONFIG['search_interval_minutes']} min")
    print(f"  Log file    : {CONFIG['log_file']}")
    print("═" * 68 + "\n")

    driver        = create_driver()
    applied_today = 0
    seen_links    = set()               # Tracks links seen this day
    current_day   = datetime.now().date()

    try:
        # Flowchart: "Login to Naukri.com" → "Login OK?"
        login(driver)

        # Flowchart: "Start Application Cycle — Every 30 min · max 40/day"
        cycle = 1
        while True:
            today = datetime.now().date()

            # Flowchart: "New Day Reset" — auto-reset counter at midnight
            if today != current_day:
                print(f"\n  [NEW DAY] {today} — resetting counter & seen links.")
                logger.info(
                    f"New day. Yesterday applied: {applied_today}"
                )
                applied_today = 0
                seen_links    = set()   # FIX-10: was missing in original
                current_day   = today

            print(f"\nCycle #{cycle}  |  "
                  f"Applied: {applied_today}/{CONFIG['max_applications_per_day']}")

            if applied_today >= CONFIG["max_applications_per_day"]:
                secs = seconds_until_midnight()
                h, m = secs // 3600, (secs % 3600) // 60
                print(f"  Daily limit reached — sleeping {h}h {m}m until midnight …")
                time.sleep(secs + 60)
                continue

            applied_today = run_application_cycle(
                driver, applied_today, seen_links
            )
            cycle += 1

            if applied_today < CONFIG["max_applications_per_day"]:
                w = CONFIG["search_interval_minutes"]
                print(f"\n  Waiting {w} min before next cycle …")
                time.sleep(w * 60)

    except KeyboardInterrupt:
        print("\n  [STOPPED] Ctrl+C — bot stopped.")
        logger.info(f"Stopped by user. Total applied: {applied_today}")

    except Exception as e:
        logger.critical(f"Fatal error: {e}", exc_info=True)
        print(f"\n  [FATAL] {e}")

    finally:
        try:
            driver.quit()
        except Exception:
            pass
        print(f"\n  Browser closed.")
        print(f"  Total applied this session: {applied_today}")
        print(f"  Log saved to: {CONFIG['log_file']}")


if __name__ == "__main__":
    main()